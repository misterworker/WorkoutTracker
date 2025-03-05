import firebase_admin
from firebase_admin import credentials, firestore
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Initialize Firebase Admin SDK
cred = credentials.Certificate('C:\\Users\\ethan\\Documents\\Work\\Impt\\'
                               'workout-wrecker-firebase-adminsdk-mw4y2-179f69fccc.json')
firebase_admin.initialize_app(cred)

# Firestore client
db = firestore.client()

bodypart_options = [
    "Shoulders", "Arms", "Legs", "Chest", "Back",
    "Full Body", "Abs", "Cardio", "Other"
]

type_options = [
    "Dumbbell", "Barbell", "Machine", "Cable",
    "Kettle Bell", "Resistance Band", "Weight Plate",
    "Calisthenics", "Other"
]

category_options = [
    "Compound", "Isolation", "Cardio", "Stretching", "None"
]

home_path = os.path.expanduser("~")
document_abs_path = os.path.join(home_path, "Documents")
print("Documents Absolute File Path: ", document_abs_path)
file_path = os.path.join(document_abs_path, f"Work\\Created Exercises.xlsx")


def create_default_exercise(doc_id, bodypart, title, type_, category):

    # Path to the collection
    doc_ref = db.collection('exercises').document(doc_id)

    # Data to store
    exercise_data = {
        'bodypart': bodypart,
        'instructions': "",
        'title': title,
        'type': type_,
        'category': category
    }

    # Set the document in Firestore
    doc_ref.create(exercise_data)
    print(f"Document {doc_id} created successfully.")

    # Log the exercise to Excel file
    log_exercise_to_excel(doc_id, bodypart, title, type_, category)


def update_all_exercises():
    if os.path.exists(file_path):
        # Load the existing workbook and read the 'Exercises' sheet
        df = pd.read_excel(file_path, sheet_name="Exercises")
        df = df.fillna("")

        # Iterate through each row of the DataFrame
        for _, row in df.iterrows():
            doc_id = row['Document ID']
            bodypart = row['Bodypart']
            instructions = row['Instructions']
            title = row['Title']
            type_ = row['Type']
            category = row['Category']


            # Update the document in Firestore
            doc_ref = db.collection('exercises').document(doc_id)

            # Data to update
            exercise_data = {
                'bodypart': bodypart,
                'instructions': instructions,
                'title': title,
                'type': type_,
                'category': category,
            }

            # Update the document in Firestore
            doc_ref.set(exercise_data)
            print(f"Document {doc_id} updated successfully.")


def log_exercise_to_excel(doc_id, bodypart, title, type_, category):
    # Create a DataFrame with the new data
    new_data = pd.DataFrame({
        'Document ID': [doc_id],
        'Bodypart': [bodypart],
        'Instructions': "",
        'Title': [title],
        'Type': [type_],
        'Category': [category],
    })

    # Check if the file exists
    if os.path.exists(file_path):
        # Load the existing workbook
        book = load_workbook(file_path)
        if "Exercises" in book.sheetnames:
            exercises_sheet = book["Exercises"]
        else:
            # Create a new sheet if it doesn't exist
            exercises_sheet = book.create_sheet("Exercises")
    else:
        # Create a new workbook and add a sheet
        from openpyxl import Workbook
        book = Workbook()
        exercises_sheet = book.active
        exercises_sheet.title = "Exercises"

        # Write the header
        header = ['Document ID', 'Bodypart', 'Instructions', 'Title', 'Type', 'Instructions Word Count', 'Category']
        exercises_sheet.append(header)

        # Write DataFrame to the sheet
    for r in dataframe_to_rows(new_data, index=False, header=False):
        exercises_sheet.append(r)

        # Save the workbook
    book.save(file_path)
    print(f"New Exercises Sheet Added and File Saved at {file_path}")


def get_input(prompt, history, options=None, backtrack_label="/BACK"):
    while True:
        value = input(prompt)
        if value.upper() == backtrack_label.upper():
            if history:
                history.pop()
            else:
                print("Nothing to backtrack.")
            return None
        if options:
            matches = [option for option in options if option.lower().startswith(value.lower())]
            if len(matches) == 1:
                value = matches[0]
                history.append(value)
                return value
            elif len(matches) > 1:
                print(f"Multiple matches found: {', '.join(matches)}. Please type more characters.")
            else:
                print(f"No match found for '{value}'. Please try again.")
        else:
            history.append(value)
            return value


def interactive_input():
    history = []
    title_df = []
    while True:
        print("\nEnter exercise details or type '/BACK' to backtrack the last input:")
        doc_id = get_input("Enter document ID ('default' will be prepended to id): ", history)
        if doc_id is None:
            continue

        # Ensure document ID starts with "default"
        if not doc_id.startswith("default_"):
            doc_id = "default_" + doc_id

        title = get_input("Enter title: ", history)
        if title is None:
            continue

        if title in title_df:
            print("WARNING: TITLE ALREADY EXISTS")

        if os.path.exists(file_path):
            df = pd.read_excel(file_path, sheet_name="Exercises")
            doc_id_df = df["Document ID"].tolist()
            title_df = df["Title"].tolist()
            if doc_id in doc_id_df:
                print("WARNING: DOCUMENT ID ALREADY EXISTS. RESUMING WILL UPDATE DATA.")

        print(bodypart_options)
        bodypart = get_input("Enter bodypart: ", history, options=bodypart_options)
        if bodypart is None:
            continue

        print(type_options)
        type_ = get_input("Enter type: ", history, options=type_options)
        if type_ is None:
            continue

        if os.path.exists(file_path):
            df = pd.read_excel(file_path, sheet_name="Exercises")
            doc_id_df = df["Document ID"].tolist()
            title_df = df["Title"].tolist()
            if doc_id in doc_id_df:
                print("WARNING: DOCUMENT ID ALREADY EXISTS. RESUMING WILL UPDATE DATA.")

        print(category_options)
        category = get_input("Enter category: ", history, options=category_options)
        if category is None:
            continue

        print("\nReview the details:")
        print(f"Document ID: {doc_id}")
        print(f"Bodypart: {bodypart}")
        print(f"Title: {title}")
        print(f"Type: {type_}")
        print(f"Category: {category}")

        confirm = input("Press Enter if details are correct: ")
        if confirm.lower() == "":
            create_default_exercise(doc_id, bodypart, title, type_, category)
            break
        else:
            print("Restarting input...\n")
            history.clear()


choice = str(input("Update, Create or Retrieve: "))
if choice == "Create":
    print("Close Create Exercise Excel File\n")
    while True:
        interactive_input()

elif choice == "Update":
    confirmation = str(input("Are you sure you want to update all exercises? (Yes to proceed...) "))
    if confirmation == "Yes":
        update_all_exercises()
